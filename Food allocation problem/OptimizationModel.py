import pyomo.environ as pyo
import pandas as pd
import datetime
import os
import math
import openpyxl
from collections import defaultdict
from itertools import combinations
from dateutil.relativedelta import relativedelta
from datetime import date
import time

######################################## INSERIMENTO DEI DATI ########################################
######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################

os.chdir(os.path.dirname(os.path.abspath(__file__)))
file_path = "prodotti_strutture_agea.xlsx"
#file_path = "prodotti_strutture_agea2.xlsx"
#file_path = "prodotti_strutture_agea3.xlsx"
#file_path = "prodotti_strutture_agea4.xlsx"
#file_path = "prodotti_strutture_agea5.xlsx"
try:
    xls = pd.ExcelFile(file_path, engine="openpyxl")
    foglio_prodotti = "prodotti"
    foglio_strutture = "strutture_caritative_uno"

    if foglio_prodotti in xls.sheet_names:
        df_prodotti = pd.read_excel(file_path, sheet_name=foglio_prodotti, dtype={"Partita": str}, engine="openpyxl")
    else:
        raise ValueError(f"Errore: Il foglio '{foglio_prodotti}' non esiste nel file.")

    if foglio_strutture in xls.sheet_names:
        df_strutture = pd.read_excel(file_path, sheet_name=foglio_strutture, engine="openpyxl")
    else:
        raise ValueError(f"Errore: Il foglio '{foglio_strutture}' non esiste nel file.")
except FileNotFoundError:
    raise FileNotFoundError("Errore: Il file non è stato trovato.")
except Exception as e:
    raise Exception(f"Si è verificato un errore durante l'importazione: {e}")

c_Pr = df_prodotti["Codice Prodotto"].tolist()
c_Pa = df_prodotti["Partita"].tolist()
prodotto_con_partita = list(zip(c_Pr, c_Pa))
df_prodotti["Data di scadenza"] = pd.to_datetime(df_prodotti["Data di scadenza"]).dt.date
data_scadenza = dict(zip(c_Pa, df_prodotti["Data di scadenza"].tolist()))
settore = dict(zip(c_Pa, df_prodotti["Settore Merceologico"].tolist())) #solo per output
campagna = dict(zip(c_Pa, df_prodotti["Campagna AGEA"].tolist())) #solo per output
quantita = dict(zip(c_Pa, df_prodotti["Quantità in colli"].tolist()))
colli_pancale = dict(zip(c_Pa, df_prodotti["Colli a pancale"].tolist()))
colli_piano = dict(zip(c_Pa, df_prodotti["Colli a piano"].tolist()))
peso = dict(zip(c_Pa, df_prodotti["Peso a collo (kg)"].tolist()))
calorie = dict(zip(c_Pa, df_prodotti["Calorie (Kcal)"].tolist()))
carboidrati = dict(zip(c_Pa, df_prodotti["Carboidrati (g)"].tolist()))
proteine = dict(zip(c_Pa, df_prodotti["Proteine (g)"].tolist()))
grassi = dict(zip(c_Pa, df_prodotti["Grassi (g)"].tolist()))
zuccheri = dict(zip(c_Pa, df_prodotti["Zuccheri (g)"].tolist()))
priorita = dict(zip(c_Pa, df_prodotti["Priorita"].tolist()))

f = defaultdict(list)
for prodotto, p in prodotto_con_partita:
    f[prodotto].append(p)

# #Lista di tuple, dove ogni tupla ha due elementi: (partita:prodotto) ('0008394908': 'A01300A')
par_to_prod = {p: c_Pr for c_Pr, p in prodotto_con_partita}

c_St = df_strutture["Codice struttura"].tolist()
assistiti = dict(zip(c_St, df_strutture["Numero Assistiti"].tolist()))
appuntamento = dict(zip(c_St, df_strutture["Codice Appuntamento"].tolist())) #solo per output
df_strutture["Freschi"] = df_strutture["Freschi"].map({'X': False, '': True}).astype(bool)
df_strutture["Infanzia"] = df_strutture["Infanzia"].map({'X': False, '': True}).astype(bool)
freschi = dict(zip(c_St, df_strutture["Freschi"].tolist()))
infanzia = dict(zip(c_St, df_strutture["Infanzia"].tolist()))
#Definisco tutti i possibili vincoli dati dalle strutture caritative sui prodotti extra infanzia e freschi che non si vogliono reperire
lista_vincoli = []
for index, row in df_strutture.iterrows():
    struttura = row['Codice struttura']
    vincoli = row['Vincoli'] 
    if pd.notna(vincoli) and str(vincoli).strip() != '':
        lista_vincoli_str = str(vincoli).split()      
        for vincolo in lista_vincoli_str:
            lista_vincoli.append((struttura, vincolo))


K = {}
for s in c_St:
    for prodotto, p in prodotto_con_partita:
        cp = colli_piano[p]
        if colli_pancale[p] % cp !=0:
             meta_pancale = colli_pancale[p] // 2 + cp
        else:
            meta_pancale = colli_pancale[p] // 2 
        K[(s, p)] = meta_pancale // cp
        if (prodotto == "C01300A"):
            K[(s, p)] = meta_pancale // cp + 1


data_corrente = date.today()
data_limite_sei_mesi = data_corrente + relativedelta(months=+6)
liste_partite_ordinate = {}
lista_prodotti_da_ordinare=set()
for prod in set(c_Pr):
    partite = f[prod]
    partite_valide = [
        par for par in partite 
        if priorita.get(par) != 1 
        and data_scadenza.get(par) >= data_corrente
    ]
    scadenza = any(
        data_scadenza.get(par) < data_limite_sei_mesi 
        for par in partite_valide
    )    
    if not scadenza:
        continue  
    partite_ordinate_lista = sorted(
        partite_valide,
        key=lambda par: (-priorita.get(par, 0), data_scadenza[par])
    )
    liste_partite_ordinate[prod] = partite_ordinate_lista
    lista_prodotti_da_ordinare.add(prod)
print(liste_partite_ordinate)

coppie_partite_ordinate = []
for lista_ordinata in liste_partite_ordinate.values():
    for i in range(len(lista_ordinata) - 1):
        partita_precedente = lista_ordinata[i]
        partita_successiva = lista_ordinata[i+1]
        coppie_partite_ordinate.append((partita_precedente, partita_successiva))

apporto_calorico_struttura = {}
richiesta = {}
for s in c_St:
    richiesta[s] = 1500 * assistiti[s] * 30
    apporto_calorico_struttura[s] = richiesta[s]

#Definisco gli input per i pesi della funzione obiettivo
w_azzeramento = float(input("Definisci il peso per l'errore nell'azzeramento della partita: "))
w_valori_nutrizionali = float(input("Definisci il peso per l'errore nel bilanciamento dei valori nutrizionali: "))
w_prodotti = float(input("Definisci il peso per l'errore nella scelta del numero di prodotti: "))
w_assegnazione_piano = float(input("Definisci il peso per l'errore nell'assegnazione di colli rispetto al dominio di x: "))
w_differenza = float(input("Definisci il peso per la differenza nella scelta del numero di prodotti tra dimensione strutture: "))
w_peso = float(input("Definisci il peso per il vincolo sul peso totale del pancale: "))

c = {}
for s in c_St:
    for prod in set(c_Pr):
        if (s, prod) in lista_vincoli:
            c[s, prod] = 0
        elif prod.startswith('I') and infanzia[s] == 0:
            c[s, prod] = 0
        elif prod[3] != '3' and freschi[s] == 0:
            c[s, prod] = 0
        else:
            c[s, prod] = 1

a = {}
for s in c_St:
    for prod in set(c_Pr):
        partite = f[prod]
        for par in partite:
            a[s, par] = c[s, prod]

PESO_RICHIESTA = {}
PESO_MASSIMO = {}
dimensione = {}
for s in c_St:
    if assistiti[s] <= 100:
        PESO_RICHIESTA[s] = 0.000001
        dimensione[s] = 1
        PESO_MASSIMO[s] = 700 * dimensione[s]
    elif assistiti[s] > 100 and assistiti[s] <= 200:
        PESO_RICHIESTA[s] = 0.0001
        dimensione[s] = 2
        PESO_MASSIMO[s] = 700 * dimensione[s]
    elif assistiti[s]> 200 and assistiti[s] <= 300:
        PESO_RICHIESTA[s] = 0.01
        dimensione[s] = 3
        PESO_MASSIMO[s] = 700 * dimensione[s]
    elif assistiti[s]> 300 and assistiti[s] <= 500:
        PESO_RICHIESTA[s] = 1
        dimensione[s] = 4
        PESO_MASSIMO[s] = 700 * dimensione[s]
    else:
        PESO_RICHIESTA[s] = 100
        dimensione[s] = 5
        PESO_MASSIMO[s] = 700 * dimensione[s]

Coppie_Valide_Singolarmente = [(s1, s2) for s1 in c_St for s2 in c_St if dimensione[s2]-dimensione[s1] == 1]

Triplette_Valide = [
    (s1, s2, prod)
    for (s1, s2) in Coppie_Valide_Singolarmente
    for prod in set(c_Pr)
    if c[s1, prod]==1 and c[s2, prod]==1
]

############################################ CREAZIONE DEL MODELLO ############################################ 
###############################################################################################################
###############################################################################################################
###############################################################################################################
###############################################################################################################

model = pyo.ConcreteModel()

model.c_Pa = pyo.Set(initialize=c_Pa)
model.c_Pr = pyo.Set(initialize=c_Pr)
model.c_St = pyo.Set(initialize=c_St)
model.tr = pyo.Set(initialize=Triplette_Valide)
model.pa_Sc = pyo.Set(initialize=coppie_partite_ordinate, dimen=2)
model.pr_Sc = pyo.Set(initialize=sorted(lista_prodotti_da_ordinare))

model.a = pyo.Param(model.c_St, model.c_Pa, initialize=a, within=pyo.Boolean)
model.c= pyo.Param(model.c_St, set(model.c_Pr), initialize=c, within=pyo.Boolean)
model.k = pyo.Param(model.c_St, model.c_Pa, initialize=K, within=pyo.NonNegativeIntegers)
model.cp = pyo.Param(model.c_Pa, initialize = colli_piano, within = pyo.NonNegativeIntegers)
model.p = pyo.Param(model.c_Pa, initialize=priorita, within=pyo.Boolean)
model.m = pyo.Param(model.c_Pa, initialize = peso, within=pyo.NonNegativeReals)
model.ca = pyo.Param(model.c_Pa, initialize = calorie, within=pyo.NonNegativeReals)
model.carb = pyo.Param(model.c_Pa, initialize = carboidrati, within=pyo.NonNegativeReals)
model.gras = pyo.Param(model.c_Pa, initialize = grassi, within=pyo.NonNegativeReals)
model.prot = pyo.Param(model.c_Pa, initialize = proteine, within=pyo.NonNegativeReals)
model.zuch = pyo.Param(model.c_Pa, initialize = zuccheri, within=pyo.NonNegativeReals)
model.app = pyo.Param(model.c_St, initialize = apporto_calorico_struttura, within=pyo.NonNegativeReals)
model.q = pyo.Param(model.c_Pa, initialize = quantita, within = pyo.NonNegativeIntegers)
model.prodmin = pyo.Param(initialize=6, within=pyo.NonNegativeIntegers)
model.prodmax = pyo.Param(initialize=10, within=pyo.NonNegativeIntegers)
def cp_hlow(model, p):
    return int(model.cp[p] // 2)

def cp_hhigh(model, p):
    return int((model.cp[p] + 1) // 2)

model.cp_hlow  = pyo.Param(model.c_Pa, initialize=cp_hlow,  within=pyo.NonNegativeIntegers)
model.cp_hhigh = pyo.Param(model.c_Pa, initialize=cp_hhigh, within=pyo.NonNegativeIntegers)

model.x = pyo.Var(model.c_St, model.c_Pa, within=pyo.NonNegativeIntegers)
model.XInt = pyo.Var(model.c_St, model.c_Pa, within=pyo.NonNegativeIntegers)
model.Xhalf = pyo.Var(model.c_St, model.c_Pa, within=pyo.Binary)
model.Hchoice = pyo.Var(model.c_St, model.c_Pa, within=pyo.Binary)

# Variabile binaria che dice se la partita p è usata per la struttura s
model.y = pyo.Var(model.c_St, model.c_Pa, within=pyo.Boolean)

#Variabile binaria che dice se la partita p è stata scelta
model.z = pyo.Var(model.c_Pa, within=pyo.Boolean)

#Variabile binaria che dice se si sta usando il prodotto p 
model.w = pyo.Var(set(model.c_Pr), within=pyo.Boolean)

model.diff = pyo.Var(model.c_St, set(model.c_Pr), within = pyo.NonNegativeIntegers)

#Variabili di scarto 
model.eps_qc = pyo.Var(model.c_Pa, within=pyo.NonNegativeIntegers)
model.eps_prod = pyo.Var(within=pyo.NonNegativeIntegers)
model.eps_car = pyo.Var( model.c_St, within=pyo.NonNegativeReals)
model.eps_gr = pyo.Var( model.c_St, within=pyo.NonNegativeReals)
model.eps_prot = pyo.Var( model.c_St, within=pyo.NonNegativeReals)
model.eps_zuc = pyo.Var( model.c_St, within=pyo.NonNegativeReals)
model.eps_ass = pyo.Var(model.c_St, model.c_Pa, within=pyo.NonNegativeIntegers)
model.eps_rich = pyo.Var(model.c_St, within=pyo.NonNegativeReals)
model.eps_diff = pyo.Var(model.tr, within=pyo.NonNegativeReals)
model.eps_peso = pyo.Var(model.c_St, within = pyo.NonNegativeReals)

#1)
def scelta_priorita_partita_rule(model, par):
    if model.p[par] == True:
        return model.z[par] == 1
    return pyo.Constraint.Skip
model.scelta_priorita_partita = pyo.Constraint(model.c_Pa, rule=scelta_priorita_partita_rule)

#2)
def prodotti_in_scadenza_rule(model, prod):
    return model.w[prod] == 1
model.prodotti_in_scadenza = pyo.Constraint(model.pr_Sc, rule=prodotti_in_scadenza_rule)

#3) 
def sequenza_scelta_rule(model, par_prec, par_succ):
    return model.z[par_succ] <= model.z[par_prec]
model.prodotto_per_data_scadenza = pyo.Constraint(model.pa_Sc, rule=sequenza_scelta_rule)

#4)
def prodotto_scelto_rule(model, par):
    prod = par_to_prod[par]
    return model.z[par] <= model.w[prod]
model.prodotto_scelto = pyo.Constraint(model.c_Pa, rule=prodotto_scelto_rule)

#5) Numero minimo e massimo di prodotti da utilizzare in una prova di assegnazione 
def numero_prodotti_min_rule(model):
    return sum(model.w[prod] for prod in set(model.c_Pr)) + model.eps_prod  >= model.prodmin
model.numero_prodotti_min = pyo.Constraint(rule = numero_prodotti_min_rule)

def numero_prodotti_max_rule(model):
    return sum(model.w[prod] for prod in set(model.c_Pr)) - model.eps_prod <= model.prodmax
model.numero_prodotti_max = pyo.Constraint(rule = numero_prodotti_max_rule)

#6) 
def prodotti_struttura_rule(model, s, prod, par):
    return model.w[prod] >= model.y[s, par]
model.prodotti_struttura = pyo.Constraint(((s, prod, par) for s in model.c_St for prod in set(model.c_Pr) for par in f[prod]), rule = prodotti_struttura_rule)

#7) Per ogni struttura può essere assegnata solo una partita di un dato prodotto
def partite_per_struttura_rule(model, s, prod):
    partite = f[prod]
    return sum(model.y[s, par] for par in partite) == model.w[prod] 
model.partite_per_struttura = pyo.Constraint(model.c_St, set(model.c_Pr), rule = partite_per_struttura_rule)

#8)
def assegnazione_piani_rule(model, s, p):
    if not model.a[s, p]:
        return model.x[s, p] == 0
    half_cap = (model.Hchoice[s,p] * model.cp_hhigh[p] + (1 - model.Hchoice[s,p]) * model.cp_hlow[p])
    return model.x[s, p] + model.eps_ass[s,p] == (model.XInt[s,p]  * model.cp[p] + model.Xhalf[s,p] * half_cap)
model.assegnazione_piani = pyo.Constraint(model.c_St, model.c_Pa, rule=assegnazione_piani_rule)

#9)
def coerenza_half_int_rule (model, s, par): 
    return model.XInt[s, par] <= model.k[s, par] * (1 - model.Xhalf[s, par])
model.coerenza_half_int = pyo.Constraint(model.c_St, model.c_Pa, rule = coerenza_half_int_rule)

#10)
def massima_assegnazione_rule(model, s, par):
    return model.x[s, par] <= model.k[s, par] * model.cp[par] * model.y[s, par] * model.a[s, par]
model.massima_assegnazione = pyo.Constraint(model.c_St, model.c_Pa, rule=massima_assegnazione_rule)

def assegnazione_minima_colli_rule(model, s, par):
    return model.x[s,par] >=  model.y[s,par] * model.a[s, par]
model.assegnazione_minima_colli = pyo.Constraint(model.c_St, model.c_Pa, rule = assegnazione_minima_colli_rule)

#11)
def richiesta_minima_rule(model, s):
     return sum(model.x[s, par] * model.m[par] * model.ca[par] * 10 for par in model.c_Pa)/model.app[s] + model.eps_rich[s] == 1
model.richiesta_minima = pyo.Constraint(model.c_St, rule=richiesta_minima_rule)

#12) Se una partita par è scelta deve essere allocata in tutte le strutture in modo che il totale rimanga 0. Se non è scelta non è allocata alcuna unità
def quantita_rule(model, par):
    return sum(model.x[s, par] for s in model.c_St) + model.eps_qc[par] == model.q[par] * model.z[par]
model.quantita = pyo.Constraint(model.c_Pa, rule = quantita_rule)

#13) Quantità limite di Carboidrati Grassi Proteine e Zuccheri
def carboidrati_min_rule(model, s):
    return sum(model.carb[par] * model.x[s,par] * model.m[par] for par in model.c_Pa)*4 + model.eps_car[s] >= sum(model.ca[par] * model.x[s,par] * model.m[par]  for par in model.c_Pa) * 0.45
model.carboidrati_min = pyo.Constraint(model.c_St, rule = carboidrati_min_rule)

def carboidrati_max_rule(model, s):
    return sum(model.carb[par] * model.x[s,par] * model.m[par] for par in model.c_Pa)*4 - model.eps_car[s] <= sum(model.ca[par] * model.x[s,par] * model.m[par]  for par in model.c_Pa) * 0.6
model.carboidrati_max = pyo.Constraint(model.c_St, rule = carboidrati_max_rule)

def grassi_min_rule(model, s):
    return sum(model.gras[par] * model.x[s,par] * model.m[par] for par in model.c_Pa)*9 + model.eps_gr[s] >= sum(model.ca[par] * model.x[s,par] * model.m[par]  for par in model.c_Pa) * 0.2
model.grassi_min = pyo.Constraint(model.c_St, rule = grassi_min_rule)

def grassi_max_rule(model, s):
    return sum(model.gras[par] * model.x[s,par] * model.m[par] for par in model.c_Pa)*9 - model.eps_gr[s] <= sum(model.ca[par] * model.x[s,par] * model.m[par]  for par in model.c_Pa) * 0.3
model.grassi_max = pyo.Constraint(model.c_St, rule = grassi_max_rule)

def proteine_min_rule(model, s):
    return sum(model.prot[par] * model.x[s,par] * model.m[par] for par in model.c_Pa)*4 + model.eps_prot[s] >= sum(model.ca[par] * model.x[s,par] * model.m[par]  for par in model.c_Pa) * 0.15
model.proteine_min = pyo.Constraint(c_St, rule = proteine_min_rule)

def proteine_max_rule(model, s):
    return sum(model.prot[par] * model.x[s,par] * model.m[par]  for par in model.c_Pa)*4 - model.eps_prot[s] <= sum(model.ca[par] * model.x[s,par] * model.m[par]  for par in model.c_Pa) * 0.2
model.proteine_max = pyo.Constraint(model.c_St, rule = proteine_max_rule)

def zuccheri_min_rule(model, s):
    return sum(model.zuch[par] * model.x[s,par] * model.m[par]  for par in model.c_Pa) - model.eps_zuc[s] <= sum(model.carb[par] * model.x[s,par] * model.m[par]  for par in model.c_Pa) * 0.25
model.zuccheri_min = pyo.Constraint(model.c_St, rule = zuccheri_min_rule)

#14) VINCOLO CHE MI DA IL VOLUME TOTALE DEL PANCALE PER LA "SODDISFAZIONE" IN TERMINI DI SENTIMENTO DI CIASCUNA STRUTTURA CARITATIVA

def valore_diff_rule(model, s, prod):
    partite = f[prod]
    return model.diff[s, prod] == sum(model.x[s, par] for par in partite)
model.valore_diff = pyo.Constraint(model.c_St, c_Pr, rule=valore_diff_rule)

def vincolo_assistiti_rule(model, s1, s2, prod): 
    return model.diff[s1, prod]  + model.eps_diff[s1, s2, prod] <= model.diff[s2, prod] 
model.vincolo_assistiti = pyo.Constraint(model.tr, rule = vincolo_assistiti_rule) 

#15) Vincolo che identifica il massimo peso che è possibile raggiungere su un pancale.
def peso_massimo_rule(model, s):
    return sum(model.x[s, par]*model.m[par] for par in model.c_Pa) - model.eps_peso[s] <= PESO_MASSIMO[s]
model.peso_massimo = pyo.Constraint(model.c_St, rule = peso_massimo_rule)

#FUNZIONE OBIETTIVO
def obiettivo_rule(model): 
    return (sum(model.eps_qc[par] for par in model.c_Pa)*w_azzeramento + (model.eps_prod)* w_prodotti  + sum(model.eps_rich[s]*PESO_RICHIESTA[s] for s in model.c_St) + sum(model.eps_ass[s, par] for s in model.c_St for par in model.c_Pa)*w_assegnazione_piano + sum( model.eps_car[s]+ model.eps_gr[s]+ model.eps_prot[s]+ model.eps_zuc[s] for s in model.c_St)*w_valori_nutrizionali + sum(model.eps_peso[s] for s in c_St)*w_peso + sum(model.eps_diff[q] for q in model.tr)*w_differenza) 
model.obiettivo = pyo.Objective(rule=obiettivo_rule, sense=pyo.minimize)

start_time = time.time()

opt = pyo.SolverFactory('gurobi')
opt.options['MIPGap'] = 0.0001 
opt.options['TimeLimit'] = 21600

result_obj = opt.solve(model, tee=True)

end_time = time.time()
execution_time = end_time - start_time

######################################## OUTPUT ########################################
########################################################################################
########################################################################################
########################################################################################
########################################################################################

#Stampa in un file di testo l'output
with open("output_modellocopia.txt", "w") as file:
    file.write(f"Tempo totale di esecuzione: {execution_time:.2f} secondi\n\n")
    totale_prodotti = sum(pyo.value(model.w[prod]) for prod in set(c_Pr))

    file.write("Pesi della funzione obiettivo\n")
    file.write(f"Azzeramento partita: {w_azzeramento}\n")
    file.write(f"Peso scelta dei prodotti: {w_prodotti}\n")
    file.write(f"Peso assegnazione in piani: {w_assegnazione_piano}\n")
    file.write(f"Peso assegnazione valori nutrizionali: {w_valori_nutrizionali}\n")
    file.write(f"Peso assegnazione differenza di dimensione struttura: {w_differenza}\n")
    file.write(f"Totale prodotti selezionati: {totale_prodotti}\n\n")

    for prod in set(c_Pr):
        partite = f[prod]
        file.write(f"\nProdotto: {prod}\n")
        for p in partite:
            valore_z = pyo.value(model.z[p])
            if valore_z == 1:
                file.write(f"  Partita {p}: z = {valore_z}\n")

    file.write("\n\n--- EPSILON GLOBALI ---\n")
    file.write(f"eps_prod = {pyo.value(model.eps_prod):.4f}\n")

    for p in c_Pa:
        if p in model.eps_qc:
            eps_val = pyo.value(model.eps_qc[p])
            if abs(eps_val) > 1e-6:
                file.write(f"eps_qc[{p}] = {eps_val:.4f}\n")

    file.write("\n\n--- RISULTATI NUTRIZIONALI PER STRUTTURA ---\n")

    for s in c_St:
        carbo = sum(carboidrati[p] * pyo.value(model.x[s, p]) * peso[p] * 10 * 4 for p in c_Pa)
        prot  = sum(proteine[p]   * pyo.value(model.x[s, p]) * peso[p] * 10 * 4 for p in c_Pa)
        grass = sum(grassi[p]     * pyo.value(model.x[s, p]) * peso[p] * 10 * 9 for p in c_Pa)
        zucch = sum(zuccheri[p]   * pyo.value(model.x[s, p]) * peso[p] * 10     for p in c_Pa)
        cal   = sum(calorie[p]    * pyo.value(model.x[s, p]) * peso[p] * 10     for p in c_Pa)
        r = richiesta[s]

        carbo_min = cal * 0.45
        carbo_max = cal * 0.6
        prot_min  = cal * 0.15
        prot_max  = cal * 0.2
        grass_min = cal * 0.2
        grass_max = cal * 0.3
        zucch_max = sum(carboidrati[p] * pyo.value(model.x[s, p] * peso[p] * 10) for p in c_Pa) * 0.25

        file.write(f"\nStruttura: {s}\n")
        file.write(f"  Peso soddisfazione richiesta: {PESO_RICHIESTA[s]} \n")
        file.write(f"  Calorie totali: {cal}, calorie richieste: {r} \n")
        file.write(f"  Carboidrati totali (kcal): {carbo_min} <= {carbo} <= {carbo_max}\n")
        file.write(f"  Proteine totali (kcal):   {prot_min} <= {prot} <= {prot_max}\n")
        file.write(f"  Grassi totali (kcal):     {grass_min} <= {grass} <= {grass_max}\n")
        file.write(f"  Zuccheri totali:          {zucch} <= {zucch_max}\n")

        note = []
        if carbo < carbo_min:  note.append("Carboidrati (difetto)")
        elif carbo > carbo_max: note.append("Carboidrati (eccesso)")

        if prot < prot_min:  note.append("Proteine (difetto)")
        elif prot > prot_max: note.append("Proteine (eccesso)")

        if grass < grass_min:  note.append("Grassi (difetto)")
        elif grass > grass_max: note.append("Grassi (eccesso)")

        if zucch > zucch_max: note.append("Zuccheri (eccesso)")

        if not note:
            file.write("Tutti i valori nutrizionali sono nei limiti.\n")
        else:
            file.write("  " + ", ".join(note) + "\n\n")

        file.write(f"Numero assistiti: {assistiti[s]} \n")

        peso_tot = sum(peso[p] * pyo.value(model.x[s, p]) for p in c_Pa)
        file.write(f"Peso totale: {peso_tot}\n")

        file.write("\n  --- Epsilon (slack) ---\n")
        file.write(f"  eps_car[{s}]   = {pyo.value(model.eps_car[s]):.4f}\n")
        file.write(f"  eps_gr[{s}]    = {pyo.value(model.eps_gr[s]):.4f}\n")
        file.write(f"  eps_prot[{s}]  = {pyo.value(model.eps_prot[s]):.4f}\n")
        file.write(f"  eps_zuc[{s}]   = {pyo.value(model.eps_zuc[s]):.4f}\n")

        eps_rich = pyo.value(model.eps_rich[s])
        if abs(eps_rich) > 1e-6:
            file.write(f"  epsilon_rich[{s}] = {eps_rich:.4f}\n")

        for p in c_Pa:
            if (s, p) in model.eps_ass:
                eps_val = pyo.value(model.eps_ass[s, p])
                if abs(eps_val) > 1e-6:
                    file.write(f"  eps_ass[{s}, {p}] = {eps_val:.4f}\n")

        file.write("\n")

    for s in c_St:
        file.write(f"Numero assistiti: {assistiti[s]} ")
        peso_tot = sum(peso[p] * pyo.value(model.x[s, p]) for p in c_Pa)
        file.write(f"Peso totale: {peso_tot}\n")


#Creazione file excel
output_data = []
for s in c_St:
    for p in c_Pa:
        val = pyo.value(model.x[s, p])
        if val > 0:
            prodotto = df_prodotti.loc[df_prodotti["Partita"] == p, "Codice Prodotto"].values[0]            
            output_data.append({
                "Struttura": s,
                "Settore Merceologico": settore[p],
                "Campagna AGEA": campagna[p],
                "Prodotto": prodotto,
                "Partita": p,
                "Quantità": val,
                "Codice Appuntamento": appuntamento[s]
            })
df_output = pd.DataFrame(output_data)
output_file = "output_modellocopia.xlsx"
df_output.to_excel(output_file, index=False) 


wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Assegnazionicopia"
partite_scelte = [par for par in c_Pa if model.z[par].value == 1]
ws.cell(row=1, column=1, value="Struttura")
for j, partita in enumerate(partite_scelte, start=2):
    ws.cell(row=1, column=j, value=partita)
for i, struttura in enumerate(c_St, start=2):
    ws.cell(row=i, column=1, value=struttura)
    for j, partita in enumerate(partite_scelte, start=2):
        qty = model.x[struttura, partita].value if (struttura, partita) in model.x else 0
        ws.cell(row=i, column=j, value=qty)
wb.save("assegnazionicopia.xlsx")